"""
Summarize and compare colorectal ViT experiment outputs.

Usage:
    python compare_experiment_outputs.py

The script reads:
    output/colorectal_exp_auto/<task>/image<size>_patch<patch>_<depth>/results_raytuned_vit_final.json

It writes compact task-specific comparison workbooks and a report:
    output/experiment_comparison/binary_comparison.xlsx
    output/experiment_comparison/multiclass_comparison.xlsx
    output/experiment_comparison/comparison_report.md

Each workbook contains separate sheets for summary, patch-size comparison,
token trends, and ViT-depth comparison.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_ROOT = BASE_DIR / "output" / "colorectal_exp_auto"
DEFAULT_OUTPUT_DIR = BASE_DIR / "output" / "experiment_comparison"
EXPERIMENT_PATTERN = re.compile(
    r"^image(?P<image_size>\d+)_patch(?P<patch_size>\d+)_(?P<vit_depth>[A-Za-z0-9_-]+)$"
)
TASKS = {"binary", "multiclass"}
TASK_ORDER = ["binary", "multiclass"]
VIT_DEPTH_ORDER = ["tiny", "small", "base"]
METRIC = "test_f1_macro"

SUMMARY_FIELDS = [
    "patch_size",
    "num_tokens",
    "vit_depth",
    "image_size",
    "experiment_id",
    "tokens_per_side",
    "model_name",
    "test_f1_macro",
    "test_accuracy",
    "test_precision_macro",
    "test_recall_macro",
    "best_val_f1_macro",
    "best_epoch",
    "roc_auc",
    "pr_auc_tumor",
    "tumor_precision",
    "tumor_recall",
    "tumor_f1",
    "lr",
    "weight_decay",
    "batch_size",
    "accum_steps",
    "effective_batch_size",
    "label_smoothing",
    "drop_rate",
    "drop_path_rate",
    "use_class_weights",
    "num_epochs",
    "task",
    "result_path",
]

TUNED_HPARAM_FIELDS = [
    "lr",
    "weight_decay",
    "batch_size",
    "accum_steps",
    "label_smoothing",
    "drop_rate",
    "drop_path_rate",
    "use_class_weights",
]

COMPARISON_FIELDS = [
    "patch_size",
    "num_tokens",
    "vit_depth",
    "image_size",
    "comparison_id",
    "experiment_id",
    "test_f1_macro",
    "test_accuracy",
    "best_val_f1_macro",
    "group_best_test_f1_macro",
    "delta_f1_from_group_best",
    "rank_within_group",
    "is_group_best",
    "tokens_same_within_group",
    "image_size_same_within_group",
    "patch_size_same_within_group",
    "vit_depth_same_within_group",
    "comparison_note",
    "group_tuned_hparams_same",
    "lr",
    "weight_decay",
    "batch_size",
    "accum_steps",
    "label_smoothing",
    "drop_rate",
    "drop_path_rate",
    "task",
]

PATCH_PAIR_FIELDS = COMPARISON_FIELDS
TOKEN_TREND_FIELDS = COMPARISON_FIELDS
VIT_DEPTH_PAIR_FIELDS = COMPARISON_FIELDS

ALWAYS_VISIBLE_FIELDS = {
    "patch_size",
    "num_tokens",
    "vit_depth",
    "image_size",
    "comparison_id",
    "experiment_id",
}

SHEET_DEFINITIONS = {
    "summary": SUMMARY_FIELDS,
    "patch_size": PATCH_PAIR_FIELDS,
    "token_trends": TOKEN_TREND_FIELDS,
    "vit_depth": VIT_DEPTH_PAIR_FIELDS,
}

SHEET_DESCRIPTIONS = {
    "summary": "One row per completed experiment, sorted by the selected metric.",
    "patch_size": "Image-size-fixed patch/token comparison. Rows with the same comparison_id share image size and ViT scale; patch size and token count change together.",
    "token_trends": "Token-count trend in long format. Rows with the same comparison_id share patch size and ViT scale while image size and token count change together.",
    "vit_depth": "ViT-depth comparison in long format. Rows with the same comparison_id share image size, patch size, and token count; compare vit_depth vertically.",
}

STALE_CSV_OUTPUTS = [
    "comparison_files.csv",
    "experiment_summary.csv",
    "patch_size_pairs.csv",
    "token_trends.csv",
    "vit_depth_pairs.csv",
    "binary_experiment_summary.csv",
    "binary_patch_size_pairs.csv",
    "binary_token_trends.csv",
    "binary_vit_depth_pairs.csv",
    "multiclass_experiment_summary.csv",
    "multiclass_patch_size_pairs.csv",
    "multiclass_token_trends.csv",
    "multiclass_vit_depth_pairs.csv",
    "experiment_comparison.xlsx",
]


Row = Dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare saved colorectal ViT experiment outputs.")
    parser.add_argument("--input-root", type=Path, default=DEFAULT_INPUT_ROOT)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--metric", default=METRIC, choices=["test_f1_macro", "test_accuracy", "best_val_f1_macro"])
    return parser.parse_args()


def read_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as file:
        value = json.load(file)
    if not isinstance(value, dict):
        raise ValueError(f"Expected JSON object: {path}")
    return value


def nested(data: Dict[str, Any], *keys: str) -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def as_number(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def collect_rows(input_root: Path) -> List[Row]:
    rows: List[Row] = []
    if not input_root.is_dir():
        raise FileNotFoundError(f"Input root does not exist: {input_root}")

    for task_dir in sorted(input_root.iterdir(), key=lambda item: item.name):
        if not task_dir.is_dir() or task_dir.name not in TASKS:
            continue
        for experiment_dir in sorted(task_dir.iterdir(), key=lambda item: item.name):
            if not experiment_dir.is_dir():
                continue
            match = EXPERIMENT_PATTERN.fullmatch(experiment_dir.name)
            if match is None:
                continue
            result_path = experiment_dir / "results_raytuned_vit_final.json"
            if not result_path.is_file():
                continue

            result = read_json(result_path)
            config = result.get("config") if isinstance(result.get("config"), dict) else {}
            image_size = int(config.get("image_size") or match.group("image_size"))
            patch_size = int(config.get("patch_size") or match.group("patch_size"))
            vit_depth = str(config.get("vit_depth") or match.group("vit_depth"))
            tokens_per_side = image_size // patch_size if patch_size else None
            num_tokens = tokens_per_side * tokens_per_side if tokens_per_side is not None else None
            batch_size = config.get("batch_size")
            accum_steps = config.get("accum_steps")

            row: Row = {
                "task": task_dir.name,
                "experiment_id": experiment_dir.name,
                "image_size": image_size,
                "patch_size": patch_size,
                "vit_depth": vit_depth,
                "tokens_per_side": tokens_per_side,
                "num_tokens": num_tokens,
                "model_name": config.get("model_name"),
                "test_f1_macro": result.get("test_f1_macro"),
                "test_accuracy": result.get("test_accuracy"),
                "test_precision_macro": result.get("test_precision_macro"),
                "test_recall_macro": result.get("test_recall_macro"),
                "best_val_f1_macro": result.get("best_val_f1_macro"),
                "best_epoch": result.get("best_epoch"),
                "roc_auc": nested(result, "test_roc_metrics", "roc_auc"),
                "pr_auc_tumor": nested(result, "test_pr_metrics", "pr_auc_tumor"),
                "tumor_precision": nested(result, "test_binary_focus_metrics", "tumor_precision"),
                "tumor_recall": nested(result, "test_binary_focus_metrics", "tumor_recall"),
                "tumor_f1": nested(result, "test_binary_focus_metrics", "tumor_f1"),
                "lr": config.get("lr"),
                "weight_decay": config.get("weight_decay"),
                "batch_size": batch_size,
                "accum_steps": accum_steps,
                "effective_batch_size": batch_size * accum_steps
                if isinstance(batch_size, int) and isinstance(accum_steps, int)
                else None,
                "label_smoothing": config.get("label_smoothing"),
                "drop_rate": config.get("drop_rate"),
                "drop_path_rate": config.get("drop_path_rate"),
                "use_class_weights": config.get("use_class_weights"),
                "num_epochs": config.get("num_epochs"),
                "result_path": str(result_path),
            }
            rows.append(row)
    return rows


def group_by(rows: Iterable[Row], keys: Sequence[str]) -> Dict[Tuple[Any, ...], List[Row]]:
    groups: Dict[Tuple[Any, ...], List[Row]] = {}
    for row in rows:
        key = tuple(row.get(name) for name in keys)
        groups.setdefault(key, []).append(row)
    return groups


def is_same_tuned_hparams(left: Row, right: Row) -> bool:
    return all(left.get(field) == right.get(field) for field in TUNED_HPARAM_FIELDS)


def depth_sort_key(row: Row) -> Tuple[int, str]:
    depth = str(row.get("vit_depth", ""))
    try:
        return (VIT_DEPTH_ORDER.index(depth), depth)
    except ValueError:
        return (len(VIT_DEPTH_ORDER), depth)


def tuned_hparams_same(rows: Sequence[Row]) -> bool:
    if len(rows) < 2:
        return True
    first = rows[0]
    return all(
        all(row.get(field) == first.get(field) for field in TUNED_HPARAM_FIELDS)
        for row in rows[1:]
    )


def same_value_within_group(group: Sequence[Row], field: str) -> bool:
    values = {normalized_value(row.get(field)) for row in group}
    return len(values) <= 1


def build_comparison_rows(
    group: Sequence[Row],
    comparison_id: str,
    sort_key: Any,
    comparison_note: str,
) -> List[Row]:
    ordered = sorted(group, key=sort_key)
    best_f1 = max(
        (float(row["test_f1_macro"]) for row in ordered if row.get("test_f1_macro") is not None),
        default=None,
    )
    same_hparams = tuned_hparams_same(ordered)
    tokens_same = same_value_within_group(ordered, "num_tokens")
    image_size_same = same_value_within_group(ordered, "image_size")
    patch_size_same = same_value_within_group(ordered, "patch_size")
    vit_depth_same = same_value_within_group(ordered, "vit_depth")
    scores = sorted(
        {float(row["test_f1_macro"]) for row in ordered if row.get("test_f1_macro") is not None},
        reverse=True,
    )
    rank_by_score = {score: index + 1 for index, score in enumerate(scores)}

    result: List[Row] = []
    for row in ordered:
        test_f1 = row.get("test_f1_macro")
        result.append(
            {
                "task": row["task"],
                "comparison_id": comparison_id,
                "patch_size": row["patch_size"],
                "num_tokens": row["num_tokens"],
                "vit_depth": row["vit_depth"],
                "image_size": row["image_size"],
                "experiment_id": row["experiment_id"],
                "test_f1_macro": test_f1,
                "test_accuracy": row.get("test_accuracy"),
                "best_val_f1_macro": row.get("best_val_f1_macro"),
                "group_best_test_f1_macro": best_f1,
                "delta_f1_from_group_best": safe_delta(test_f1, best_f1),
                "rank_within_group": rank_by_score.get(float(test_f1)) if test_f1 is not None else None,
                "is_group_best": test_f1 is not None and best_f1 is not None and float(test_f1) == best_f1,
                "tokens_same_within_group": tokens_same,
                "image_size_same_within_group": image_size_same,
                "patch_size_same_within_group": patch_size_same,
                "vit_depth_same_within_group": vit_depth_same,
                "comparison_note": comparison_note,
                "group_tuned_hparams_same": same_hparams,
                "lr": row.get("lr"),
                "weight_decay": row.get("weight_decay"),
                "batch_size": row.get("batch_size"),
                "accum_steps": row.get("accum_steps"),
                "label_smoothing": row.get("label_smoothing"),
                "drop_rate": row.get("drop_rate"),
                "drop_path_rate": row.get("drop_path_rate"),
            }
        )
    return result


def build_patch_pairs(rows: Sequence[Row]) -> List[Row]:
    comparison_rows: List[Row] = []
    groups = group_by(rows, ["task", "image_size", "vit_depth"])
    for (task, image_size, vit_depth), group in sorted(groups.items()):
        by_patch = {row["patch_size"]: row for row in group}
        if 16 not in by_patch or 32 not in by_patch:
            continue
        pair = [by_patch[16], by_patch[32]]
        comparison_id = f"image{image_size}_depth{vit_depth}"
        comparison_rows.extend(
            build_comparison_rows(
                pair,
                comparison_id,
                lambda row: row["patch_size"],
                "image_size and vit_depth are fixed; patch_size and num_tokens change together, so this is not a pure patch-size ablation.",
            )
        )
    return comparison_rows


def build_token_trends(rows: Sequence[Row]) -> List[Row]:
    trend_rows: List[Row] = []
    groups = group_by(rows, ["task", "patch_size", "vit_depth"])
    for (task, patch_size, vit_depth), group in sorted(groups.items()):
        if len(group) < 2:
            continue
        comparison_id = f"patch{patch_size}_depth{vit_depth}"
        trend_rows.extend(
            build_comparison_rows(
                group,
                comparison_id,
                lambda row: (row["num_tokens"], row["image_size"]),
                "patch_size and vit_depth are fixed; image_size and num_tokens change together, so this is not a pure token-count ablation.",
            )
        )
    return trend_rows


def build_vit_depth_pairs(rows: Sequence[Row]) -> List[Row]:
    comparison_rows: List[Row] = []
    groups = group_by(rows, ["task", "image_size", "patch_size"])
    for (task, image_size, patch_size), group in sorted(groups.items()):
        available_depths = {row["vit_depth"] for row in group}
        ordered_depths = [depth for depth in VIT_DEPTH_ORDER if depth in available_depths]
        if len(ordered_depths) < 2:
            continue
        comparable = [row for row in group if row["vit_depth"] in ordered_depths]
        comparison_id = f"image{image_size}_patch{patch_size}"
        comparison_rows.extend(
            build_comparison_rows(
                comparable,
                comparison_id,
                depth_sort_key,
                "image_size, patch_size, and num_tokens are fixed; vit_depth changes. Ray Tune hyperparameters may still differ.",
            )
        )
    return comparison_rows

def safe_delta(left: Any, right: Any) -> Any:
    if left is None or right is None:
        return None
    return float(left) - float(right)


def sort_by_metric(rows: Sequence[Row], metric: str) -> List[Row]:
    return sorted(
        rows,
        key=lambda row: (
            row.get(metric) is not None,
            as_number(row.get(metric)) if row.get(metric) is not None else -1,
        ),
        reverse=True,
    )


def format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.4f}"
    return str(value)


def markdown_table(rows: Sequence[Row], fields: Sequence[str], limit: int | None = None) -> str:
    selected = list(rows[:limit] if limit is not None else rows)
    if not selected:
        return "_No rows._"
    header = "| " + " | ".join(fields) + " |"
    divider = "| " + " | ".join(["---"] * len(fields)) + " |"
    body = [
        "| " + " | ".join(format_value(row.get(field)) for field in fields) + " |"
        for row in selected
    ]
    return "\n".join([header, divider, *body])


def mean(values: Iterable[Any]) -> Any:
    numeric = [float(value) for value in values if value is not None]
    if not numeric:
        return None
    return sum(numeric) / len(numeric)


def task_title(task: str) -> str:
    return "Binary" if task == "binary" else "Multiclass"


def cleanup_stale_outputs(out_dir: Path) -> List[str]:
    removed: List[str] = []
    for name in STALE_CSV_OUTPUTS:
        path = out_dir / name
        if not path.is_file():
            continue
        try:
            path.unlink()
            removed.append(name)
        except OSError as error:
            print(f"Warning: could not remove stale output {path}: {error}")
    return removed


def excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def normalized_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def constant_fields(rows: Sequence[Row], fields: Sequence[str]) -> Dict[str, Any]:
    constants: Dict[str, Any] = {}
    if not rows:
        return constants
    for field in fields:
        values = [normalized_value(row.get(field)) for row in rows]
        if all(value == values[0] for value in values):
            constants[field] = rows[0].get(field)
    return constants


def find_fixed_fields(rows: Sequence[Row], fields: Sequence[str]) -> Dict[str, Any]:
    fixed: Dict[str, Any] = {}
    for field, value in constant_fields(rows, fields).items():
        if field in ALWAYS_VISIBLE_FIELDS:
            continue
        if normalized_value(value) == "":
            continue
        fixed[field] = value
    return fixed


def variable_fields(rows: Sequence[Row], fields: Sequence[str]) -> List[str]:
    constants = constant_fields(rows, fields)
    return [field for field in fields if field in ALWAYS_VISIBLE_FIELDS or field not in constants]


def fixed_summary(fixed: Dict[str, Any]) -> str:
    if not fixed:
        return "None"
    return "; ".join(f"{field}={format_value(value)}" for field, value in fixed.items())


def append_report_table(
    lines: List[str],
    rows: Sequence[Row],
    table_fields: Sequence[str],
    fixed_fields: Sequence[str] | None = None,
    limit: int | None = None,
) -> None:
    display_fields = variable_fields(rows, table_fields)
    if not display_fields and table_fields:
        display_fields = [table_fields[0]]
    lines.append(markdown_table(rows, display_fields, limit=limit))
    fixed = find_fixed_fields(rows, fixed_fields or table_fields)
    if fixed:
        lines.append("")
        lines.append(f"Fixed variables: {fixed_summary(fixed)}")


def append_table(worksheet: Any, rows: Sequence[Row], fields: Sequence[str], header_fill: str = "D9EAF7") -> None:
    worksheet.append(list(fields))
    for cell in worksheet[worksheet.max_row]:
        cell.font = worksheet.parent._named_styles[0].font.copy(bold=True)
        cell.fill = worksheet.parent._header_fill if hasattr(worksheet.parent, "_header_fill") else cell.fill
    for row in rows:
        worksheet.append([excel_value(row.get(field)) for field in fields])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def format_sheet(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        header = str(column_cells[0].value or "")
        max_len = len(header)
        for cell in column_cells[1:101]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 42)


def write_excel_workbook(path: Path, sheets: Dict[str, Tuple[Sequence[Row], Sequence[str]]]) -> None:
    try:
        from copy import copy
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as error:
        raise RuntimeError("openpyxl is required to write comparison workbooks") from error

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    notes = workbook.create_sheet(title="notes")
    notes.append(["sheet", "comparison", "fixed_variables_removed_from_table"])
    for cell in notes[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for sheet_name, (rows, fields) in sheets.items():
        fixed = find_fixed_fields(rows, fields)
        notes.append([sheet_name, SHEET_DESCRIPTIONS.get(sheet_name, ""), fixed_summary(fixed)])
    format_sheet(notes)

    for sheet_name, (rows, fields) in sheets.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        display_fields = variable_fields(rows, fields)
        if not display_fields:
            display_fields = [fields[0]] if fields else []
        worksheet.append(list(display_fields))
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = copy(header_fill)
        for row in rows:
            worksheet.append([excel_value(row.get(field)) for field in display_fields])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        format_sheet(worksheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_report(
    path: Path,
    rows_by_task: Dict[str, Sequence[Row]],
    patch_pairs_by_task: Dict[str, Sequence[Row]],
    token_trends_by_task: Dict[str, Sequence[Row]],
    vit_depth_pairs_by_task: Dict[str, Sequence[Row]],
    metric: str,
) -> None:
    lines: List[str] = []
    lines.append("# Experiment Comparison Report")
    lines.append("")
    lines.append("## How to read this")
    lines.append("")
    lines.append("- All comparisons are task-separated. Binary and multiclass are not compared against each other.")
    lines.append("- Main spreadsheets: `binary_comparison.xlsx` and `multiclass_comparison.xlsx`.")
    lines.append("- In each spreadsheet, open `notes` first for hidden fixed variables, then use `summary`, `patch_size`, `token_trends`, and `vit_depth` sheets.")
    lines.append("- Comparison sheets use long format: one row is one experiment; rows with the same `comparison_id` are compared together.")
    lines.append("- `patch_size` fixes image size and ViT scale; patch size and token count change together, so it is not a pure patch-size ablation.")
    lines.append("- `token_trends` fixes patch size and ViT scale; image size and token count change together, so it is not a pure token-count ablation.")
    lines.append("- `vit_depth` fixes image size, patch size, and token count; only ViT scale changes, though Ray Tune hyperparameters may still differ.")
    lines.append("- Because Ray Tune searched hyperparameters separately for each architecture, these are best-result comparisons, not strict one-variable ablations.")
    lines.append("")

    total = sum(len(rows_by_task.get(task, [])) for task in TASK_ORDER)
    lines.append(f"Finished experiments found: {total}")
    for task in TASK_ORDER:
        lines.append(f"- {task}: {len(rows_by_task.get(task, []))}")
    lines.append("")

    for task in TASK_ORDER:
        task_rows = list(rows_by_task.get(task, []))
        patch_pairs = list(patch_pairs_by_task.get(task, []))
        token_trends = list(token_trends_by_task.get(task, []))
        vit_depth_pairs = list(vit_depth_pairs_by_task.get(task, []))

        lines.append(f"## {task_title(task)}")
        lines.append("")
        lines.append("### Top Results")
        lines.append("")
        append_report_table(
            lines,
            sort_by_metric(task_rows, metric),
            [
                "patch_size",
                "num_tokens",
                "vit_depth",
                "image_size",
                "experiment_id",
                "test_f1_macro",
                "test_accuracy",
                "best_val_f1_macro",
                "lr",
                "batch_size",
                "drop_rate",
            ],
            SUMMARY_FIELDS,
            limit=12,
        )
        lines.append("")

        lines.append("### Patch Size Comparison")
        lines.append("")
        append_report_table(
            lines,
            patch_pairs,
            [
                "patch_size",
                "num_tokens",
                "vit_depth",
                "image_size",
                "comparison_id",
                "experiment_id",
                "test_f1_macro",
                "test_accuracy",
                "delta_f1_from_group_best",
                "rank_within_group",
                "is_group_best",
                "tokens_same_within_group",
                "comparison_note",
            ],
            PATCH_PAIR_FIELDS,
        )
        lines.append("")
        lines.append("Read this by grouping rows with the same `comparison_id`; lower `rank_within_group` is better.")
        lines.append("")

        lines.append("### Token Trends")
        lines.append("")
        append_report_table(
            lines,
            token_trends,
            [
                "patch_size",
                "num_tokens",
                "vit_depth",
                "image_size",
                "comparison_id",
                "experiment_id",
                "test_f1_macro",
                "test_accuracy",
                "delta_f1_from_group_best",
                "rank_within_group",
                "is_group_best",
                "tokens_same_within_group",
                "comparison_note",
            ],
            TOKEN_TREND_FIELDS,
        )
        lines.append("")
        lines.append("Read this by grouping rows with the same `comparison_id`; token count changes with image size and patch size.")
        lines.append("")

        lines.append("### ViT Depth Comparison")
        lines.append("")
        append_report_table(
            lines,
            vit_depth_pairs,
            [
                "patch_size",
                "num_tokens",
                "vit_depth",
                "image_size",
                "comparison_id",
                "experiment_id",
                "test_f1_macro",
                "test_accuracy",
                "delta_f1_from_group_best",
                "rank_within_group",
                "is_group_best",
                "tokens_same_within_group",
                "comparison_note",
            ],
            VIT_DEPTH_PAIR_FIELDS,
        )
        lines.append("")
        lines.append("Read this by grouping rows with the same `comparison_id`; lower `rank_within_group` is better.")
        lines.append("")

    lines.append("## Strict Ablation Recommendation")
    lines.append("")
    lines.append(
        "For a strict claim such as 'patch size alone caused the change' or 'ViT scale alone caused the change', "
        "rerun final training with a fixed config and change only that variable, or compare repeated runs with "
        "the same random seeds and report mean/std. Use these generated pair tables as exploratory evidence, "
        "not as controlled ablations by themselves."
    )
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")

def main() -> None:
    args = parse_args()
    rows = collect_rows(args.input_root)
    args.out_dir.mkdir(parents=True, exist_ok=True)

    rows_by_task: Dict[str, List[Row]] = {
        task: [row for row in rows if row["task"] == task]
        for task in TASK_ORDER
    }
    patch_pairs_by_task: Dict[str, List[Row]] = {}
    token_trends_by_task: Dict[str, List[Row]] = {}
    vit_depth_pairs_by_task: Dict[str, List[Row]] = {}
    workbook_paths: List[Path] = []

    for task in TASK_ORDER:
        task_rows = rows_by_task[task]
        patch_pairs = build_patch_pairs(task_rows)
        token_trends = build_token_trends(task_rows)
        vit_depth_pairs = build_vit_depth_pairs(task_rows)

        patch_pairs_by_task[task] = patch_pairs
        token_trends_by_task[task] = token_trends
        vit_depth_pairs_by_task[task] = vit_depth_pairs

        sheets = {
            "summary": (sort_by_metric(task_rows, args.metric), SUMMARY_FIELDS),
            "patch_size": (patch_pairs, PATCH_PAIR_FIELDS),
            "token_trends": (token_trends, TOKEN_TREND_FIELDS),
            "vit_depth": (vit_depth_pairs, VIT_DEPTH_PAIR_FIELDS),
        }
        workbook_path = args.out_dir / f"{task}_comparison.xlsx"
        write_excel_workbook(workbook_path, sheets)
        workbook_paths.append(workbook_path)

    removed_outputs = cleanup_stale_outputs(args.out_dir)
    write_report(
        args.out_dir / "comparison_report.md",
        rows_by_task,
        patch_pairs_by_task,
        token_trends_by_task,
        vit_depth_pairs_by_task,
        args.metric,
    )

    print(f"Found experiments: {len(rows)}")
    for workbook_path in workbook_paths:
        print(f"Wrote: {workbook_path}")
    print(f"Wrote: {args.out_dir / 'comparison_report.md'}")
    if removed_outputs:
        print(f"Removed stale output files: {len(removed_outputs)}")

if __name__ == "__main__":
    main()
